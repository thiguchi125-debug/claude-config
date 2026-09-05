#!/usr/bin/env python3
"""市民要望管理台帳 xlsx を生成（openpyxl）。Drive同期フォルダへ置き、草川が「Googleスプレッドシートとして保存」で native 化する。"""
import sys, json, datetime, os
sys.path.insert(0, os.path.dirname(__file__))
from migration_rows import R
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter as L

TD = json.load(open('/private/tmp/claude-501/-Users-kusakawatakuya/27f98dab-3eb2-4056-8895-6f8ebf1e751e/scratchpad/td_all.json'))
CREATED = {t['id']: (t['created'] or '')[:10] for t in TD}

COLS = [  # (見出し, 幅)
 ("No",8),("受付日",11),("地区",10),("経路",12),("要望内容（件名）",40),("詳細・経緯",48),
 ("相談者",14),("連絡手段",12),("分類",14),("相手先（担当課など）",22),("TEL",13),
 ("照会日",11),("回答・進捗メモ",40),("回答日",11),("ステータス",13),("次アクション",34),
 ("次アクション期限",12),("相談者へ報告日",12),("最終更新",11),("滞留日数",9),("要対応",7),
 ("関連リンク",22),("旧TodoistID",26),
]
H = {c[0]: i+1 for i, c in enumerate(COLS)}
def col(name): return L(H[name])

STATUS = ["受付","照会中","回答済・報告前","保留","完了"]
BUNRUI = ["道路・交通安全","河川・水路","公園・施設","防犯・防災","環境・衛生","教育・学校","福祉・医療","農林・獣害","交通・バス","その他"]
KEIRO = ["口頭","電話","報告会","ご意見箱","SNS DM","Instagram DM","LINE","メール","自治会","市民意見","市民の声","Discord投げ込み","Todoist移行","その他"]

wb = Workbook()
ws = wb.active; ws.title = "台帳"
thin = Side(style="thin", color="CCCCCC"); border = Border(left=thin,right=thin,top=thin,bottom=thin)
hfill = PatternFill("solid", fgColor="1F4E3D"); hfont = Font(bold=True, color="FFFFFF")
for i,(name,w) in enumerate(COLS,1):
    c = ws.cell(row=1, column=i, value=name); c.fill=hfill; c.font=hfont
    c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); c.border=border
    ws.column_dimensions[L(i)].width = w
ws.row_dimensions[1].height = 30
ws.freeze_panes = "F2"

NROWS = 120  # slim: Drive MCP upload用
today = datetime.date.today()
for n, r in enumerate(R, 1):
    row = n+1
    rec = CREATED.get(r['ids'][0], '') if r.get('ids') else ''
    vals = {
      "No": f"KY-{n:03d}", "受付日": rec, "地区": r.get('k',''), "経路": r.get('r',''),
      "要望内容（件名）": r['t'], "詳細・経緯": r.get('d',''), "相談者": r.get('s',''), "連絡手段": r.get('c',''),
      "分類": r.get('b',''), "相手先（担当課など）": r.get('a',''), "TEL": r.get('l',''),
      "ステータス": r.get('st','受付'), "次アクション": r.get('na',''), "次アクション期限": r.get('due',''),
      "最終更新": today.isoformat(), "関連リンク": r.get('link',''), "旧TodoistID": ",".join(r.get('ids',[])),
    }
    for k,v in vals.items():
        if v not in ("", None):
            c = ws.cell(row=row, column=H[k], value=v)
            if k in ("受付日","次アクション期限","最終更新") and len(str(v))==10:
                c.value = datetime.date.fromisoformat(v); c.number_format="yyyy-mm-dd"
for row in range(2, NROWS+1):
    r_ = str(row)
    ws[f"{col('滞留日数')}{r_}"] = (f'=IF(OR({col("No")}{r_}="",{col("ステータス")}{r_}="完了"),"",'
        f'TODAY()-MAX({col("受付日")}{r_},{col("照会日")}{r_},{col("回答日")}{r_},{col("最終更新")}{r_}))')
    ws[f"{col('要対応')}{r_}"] = (f'=IF({col("No")}{r_}="","",IF(AND({col("ステータス")}{r_}<>"完了",{col("ステータス")}{r_}<>"保留",'
        f'OR(N({col("滞留日数")}{r_})>=14,AND({col("次アクション期限")}{r_}<>"",{col("次アクション期限")}{r_}<=TODAY()),'
        f'{col("ステータス")}{r_}="回答済・報告前")),"⚠",""))')
    for k in ("受付日","照会日","回答日","次アクション期限","相談者へ報告日","最終更新"):
        ws[f"{col(k)}{r_}"].number_format = "yyyy-mm-dd"
    for c in range(1, len(COLS)+1):
        cell = ws.cell(row=row, column=c)
        cell.alignment = Alignment(vertical="top", wrap_text=(c in (H["要望内容（件名）"],H["詳細・経緯"],H["回答・進捗メモ"],H["次アクション"])))

def dv(formula, rng, **kw):
    d = DataValidation(type="list", formula1=formula, allow_blank=True, **kw); ws.add_data_validation(d); d.add(rng)
dv('"'+",".join(STATUS)+'"', f"{col('ステータス')}2:{col('ステータス')}{NROWS}")
dv('"'+",".join(BUNRUI)+'"', f"{col('分類')}2:{col('分類')}{NROWS}")
dv('"'+",".join(KEIRO)+'"', f"{col('経路')}2:{col('経路')}{NROWS}")
dv("=相手先マスタ!$A$2:$A$60", f"{col('相手先（担当課など）')}2:{col('相手先（担当課など）')}{NROWS}", showErrorMessage=False)

rng = f"A2:{L(len(COLS))}{NROWS}"
S = f"${col('ステータス')}2"; Y = f"${col('要対応')}2"
ws.conditional_formatting.add(rng, FormulaRule(formula=[f'{S}="完了"'], fill=PatternFill("solid", fgColor="E0E0E0"), font=Font(color="808080")))
ws.conditional_formatting.add(rng, FormulaRule(formula=[f'{S}="保留"'], fill=PatternFill("solid", fgColor="DDEBF7")))
ws.conditional_formatting.add(rng, FormulaRule(formula=[f'{S}="回答済・報告前"'], fill=PatternFill("solid", fgColor="FFF2CC")))
ws.conditional_formatting.add(rng, FormulaRule(formula=[f'{S}="照会中"'], fill=PatternFill("solid", fgColor="F2F7EC")))
ws.conditional_formatting.add(rng, FormulaRule(formula=[f'{Y}="⚠"'], font=Font(color="C00000", bold=True)))
ws.auto_filter.ref = f"A1:{L(len(COLS))}{NROWS}"

# --- 相手先マスタ ---
m = wb.create_sheet("相手先マスタ")
mh = ["相手先（台帳の表記）","部","TEL","メール","所管・メモ","確認日"]
for i,h in enumerate(mh,1):
    c=m.cell(row=1,column=i,value=h); c.fill=hfill; c.font=hfont; c.border=border
MASTER = [
 ("建設管理課 道路保全G","建設部","0595-84-5041","hozen@city.kameyama.mie.jp","既存市道の草刈り・支障木・舗装・側溝の維持保全。修繕要望は原則自治会長経由、急を要する陥没等は直接連絡可","2026-08-31"),
 ("土木課","建設部","","","道路整備G・河川流域G（新設改良側）",""),
 ("都市整備課","建設部","","","",""),
 ("建築住宅課","建設部","","","空き家関連（要確認）",""),
 ("学校教育課","教育委員会","0595-84-5072","","通学路・いじめ対応・登下校基準",""),
 ("教育総務課","教育委員会","","","学校施設（体育館・網戸・駐車場）",""),
 ("教育委員会","教育委員会","","","",""),
 ("消防本部（消防総務課）","消防本部","","","AED・消火栓・消防団",""),
 ("防災安全課","","","","防犯灯・防犯カメラ・防災",""),
 ("商工観光課","","","","観光・地域ブランドG",""),
 ("文化課","","","","多文化共生",""),
 ("地域福祉課 高齢者支援G","健康福祉部","","","補聴器助成",""),
 ("子ども未来課（要確認）","","","","保育園（課名は市公式で確認）",""),
 ("農林振興課（要確認）","","","","獣害（課名は市公式で確認）",""),
 ("農林（田中主幹）","","","","神向谷マンボ",""),
 ("亀山市社会福祉協議会","（外部）","","","",""),
 ("三重県鈴鹿建設事務所","（県）","","","県道・県管理河川",""),
 ("河川管理者（三重県）","（県）","","","",""),
 ("亀山警察署","（外部）","","","信号・横断歩道・速度対策は公安委員会ルート",""),
 ("自治会","（地域）","","","自治会長経由の要望書",""),
 ("市担当課","","","","担当課未特定のときの仮置き。特定後に書き換える",""),
]
for r_i, rec in enumerate(MASTER, 2):
    for c_i, v in enumerate(rec, 1):
        cell = m.cell(row=r_i, column=c_i, value=v); cell.border=border
for i,w in enumerate([28,14,14,30,60,12],1): m.column_dimensions[L(i)].width=w
m.freeze_panes="A2"

# --- 集計 ---
g = wb.create_sheet("集計")
g["A1"]="市民要望 集計ダッシュボード"; g["A1"].font=Font(bold=True,size=14)
g["A2"]="※台帳に行を足す・ステータスを変えると自動更新"
g["A4"]="ステータス"; g["B4"]="件数"
for i,s in enumerate(STATUS,5):
    g[f"A{i}"]=s; g[f"B{i}"]=f'=COUNTIF(台帳!{col("ステータス")}:{col("ステータス")},A{i})'
g["A11"]="要対応（⚠）"; g["B11"]=f'=COUNTIF(台帳!{col("要対応")}:{col("要対応")},"⚠")'
g["A12"]="未完了 合計"; g["B12"]='=B5+B6+B7+B8'
g["D4"]="分類"; g["E4"]="未完了件数"
for i,b in enumerate(BUNRUI,5):
    g[f"D{i}"]=b; g[f"E{i}"]=f'=COUNTIFS(台帳!{col("分類")}:{col("分類")},D{i},台帳!{col("ステータス")}:{col("ステータス")},"<>完了")'
g["G4"]="相手先"; g["H4"]="未完了件数"
for i,rec in enumerate(MASTER,5):
    g[f"G{i}"]=rec[0]; g[f"H{i}"]=f'=COUNTIFS(台帳!{col("相手先（担当課など）")}:{col("相手先（担当課など）")},G{i},台帳!{col("ステータス")}:{col("ステータス")},"<>完了")'
g["A16"]="⚠ 要対応 一覧（滞留14日以上／期限超過／報告前）"; g["A16"].font=Font(bold=True)
g["A17"]=(f'=IFERROR(FILTER({{台帳!{col("No")}2:{col("No")},台帳!{col("要望内容（件名）")}2:{col("要望内容（件名）")},台帳!{col("相手先（担当課など）")}2:{col("相手先（担当課など）")},'
          f'台帳!{col("ステータス")}2:{col("ステータス")},台帳!{col("次アクション")}2:{col("次アクション")},台帳!{col("滞留日数")}2:{col("滞留日数")}}},'
          f'台帳!{col("要対応")}2:{col("要対応")}="⚠"),"なし")')
for hc in ("A4","B4","D4","E4","G4","H4"): g[hc].fill=hfill; g[hc].font=hfont
for i,w in enumerate([22,10,4,18,12,4,28,12],1): g.column_dimensions[L(i)].width=w

# --- 使い方 ---
u = wb.create_sheet("使い方")
lines = [
 "市民要望管理台帳（市役所対応の市民要望はTodoistでなくこのシートで管理・2026-09-05〜）",
 "",
 "■ 1行＝1案件。受付→担当課へ照会→回答→相談者へ報告→完了 を同じ行で追う。",
 "■ ステータス：受付（未着手）／照会中（相手に投げて返事待ち）／回答済・報告前（返事は来た・相談者にまだ伝えていない）／保留／完了",
 "■ 要対応⚠ は自動：滞留14日以上・次アクション期限を過ぎた・回答済なのに相談者へ未報告 のどれかで点く（完了・保留は除外）。",
 "■ 滞留日数＝今日 − 直近の動き（受付日・照会日・回答日・最終更新の最大）。動いたら最終更新を今日にする（Apps Script導入後は自動）。",
 "■ 相手先はプルダウン（相手先マスタ）。無い課はマスタに1行足す。課名は市公式サイトで確認してから（「道路河川課」は存在しない）。",
 "■ 相談者の氏名・連絡先はこのシートにだけ書く（置き場：00_名簿・個人情報／grep・発信の対象外）。Todoist・Notion・SNSには転記しない。",
 "■ 集計タブ：ステータス別／分類別／相手先別の件数と、⚠一覧。朝のブリーフィング（ohayo）はこの⚠一覧を読む。",
 "■ Claude側からの追加・更新は Apps Script Web App 経由（手順書：~/outputs/yobo-sheet/SETUP_AppsScript.md）。",
 "■ 旧TodoistID：2026-09-05の移行元。移行完了後は参照用（Todoist側は完了処理済み）。",
 "■ 議会の一般質問ネタ・政策調査・イベント・選挙のタスクは引き続きTodoist。ここは「市民から頼まれて役所（県・警察・自治会含む）と動く案件」だけ。",
]
for i,t in enumerate(lines,1): u[f"A{i}"]=t
u["A1"].font=Font(bold=True,size=13); u.column_dimensions["A"].width=120

out_dir = os.path.expanduser("~/outputs/yobo-sheet"); os.makedirs(out_dir, exist_ok=True)
out = os.path.join(out_dir, "市民要望管理台帳_v1_slim.xlsx"); wb.save(out); print("saved", out, len(R), "rows")
